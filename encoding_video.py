#!/usr/bin/env python3
import os
import sys
import argparse
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import json
import math
import psutil
import glob
import time
import subprocess
import numpy as np
import threading
from codecarbon import OfflineEmissionsTracker
import csv
np.set_printoptions(suppress=True)
from sty import fg, bg, ef, rs

"""
usage :

    Installation : 
        pip3 install -r requirements.txt
        sudo chmod a=r /sys/class/powercap/intel-rapl:0/energy_uj 
  
    cp encoding_info_empty.xlsx  encoding_info.xlsx
    
    python3 /your-path-to-the-project/encoding_video.py
	    -s /your-path-to-the-project/original_dataset \
	    -r '/your-path-to-the-project/encoding_results.xlsx' \
	    -m psnr ssim vmaf
	   
	python3 '/media/ridha/D81821F01821CE76/Doctorat/Projet4_Energy/Project_energy_final/CTC/code/encoding_video.py' \
	    -s '/media/ridha/D81821F01821CE76/Doctorat/Projet4_Energy/Project_energy_final/CTC/code/original_dataset' \
	    -r '/media/ridha/D81821F01821CE76/Doctorat/Projet4_Energy/Project_energy_final/CTC/code/encoding_results.xlsx' \
	    -m psnr ssim vmaf

"""

#..............................................................................................
class video_encoding:
        
    def __init__(self, results, path_envid):
        self.results = results
        self.path_envid = path_envid
        self.project_path =os.path.dirname(self.path_envid)
        self.info_sys = []
        self.info_sys_all = []
        self.info_sys_gpu = []
        self.info_sys_all_gpu = []
        self.stop_threads = False 
        
    def find_index_row(self, sheet, title_row):
        for i in range(1, sheet.max_row + 1):
            cell_name = sheet.cell(row = i, column = 1).value
            if title_row == cell_name:
                return i  
    
    # decorater used to block function printing to the console
    def blockPrinting(self, func):
        def func_wrapper(*args, **kwargs):
            # block all printing to the console
            sys.stdout = open(os.devnull, 'w')
            # call the method in question
            value = func(*args, **kwargs)
            # enable all printing to the console
            sys.stdout = sys.__stdout__
            # pass the return value of the method back
            return value

        return func_wrapper      
    
    def average_column_matrix(self, matrix):
        average = []
        for j in range(len(matrix[0])) :
            count = 0
            sumv = 0
            ave = 0
            for i in range(len(matrix)):
                if matrix[i][j] != 0.0 :
                    count = count+1
                    sumv = sumv + matrix[i][j]
            if count != 0:
                ave  = sumv / float(count)
            average.append(ave)
            #print('average(', matrix[:, j], ') =', ave)

        return average

    def get_pid(self, process_name):
        try:
            l = [int(i) for i in subprocess.check_output(["pidof",process_name]).split()] 
        except subprocess.CalledProcessError as e:
            l = []
            #print("The process does not exist")
        return l

    
    
    def cpu_mem_usage_all(self, proc_name, interval):
        while True:
            pids = self.get_pid(proc_name)
            if pids:
                try:
                    self.info_sys = []
                    proc = psutil.Process(pids[0])
                    #self.info_pid.append(float(pids[0]))
                    self.info_sys.append(float(psutil.cpu_percent(interval)))
                    self.info_sys.append(float(proc.cpu_percent(interval)))
                        
                    self.info_sys.append(float(psutil.virtual_memory().total) / float(1024 ** 2)) 
                    self.info_sys.append(float(psutil.virtual_memory().available) / float(1024 ** 2)) 
                    self.info_sys.append(float(psutil.virtual_memory().free) / float(1024 ** 2)) 
                    self.info_sys.append(float(psutil.virtual_memory().used) / float(1024 ** 2))
                    self.info_sys.append(float(psutil.virtual_memory().percent))
                        
                    self.info_sys.append(proc.memory_info().rss / float(1024 ** 2))
                    self.info_sys.append(float(proc.memory_percent()))
                    
                    #load1 = psutil.getloadavg()
                    #for la in load1:
                    #    self.info_sys.append(float(la))
                    
                    self.info_sys_all.append(self.info_sys)      
                    
                except psutil.NoSuchProcess as e:
                    inor=1
                    #print(e)
            if self.stop_threads:
                break
	
    def encoding(self, metric, q, wb0, wb1, wb2):
       
        VVenC_path = glob.glob('**/bin/release-static/vvencapp', recursive=True)
        if VVenC_path == []:
            print('Please install VVenC')
            exit()
        VVdeC_path = glob.glob('**/bin/release-static/vvdecapp', recursive=True)
        if VVdeC_path ==  []:
            print('Please install VVdeC')
            exit()
        
        sheet0 = wb0.worksheets[0]
        sheet1 = wb1.worksheets[0]
        sheet2 = wb2.worksheets[0]
        self.stop_threads = False
        proc_name = 'ffmpeg'
        #'libx265', 
        codec_list_cpu = ['libx264', 'libx265', 'libvpx-vp9', 'VVenC', 'libsvtav1']
        
        header1 = "|{:^50}|{:^10}|{:^13}|{:^14}|{:^13}|{:^8}|{:^8}|{:^8}|".format("Video", "Time (s)", "Bitrate(kb/s)", "Energy (Wh)", "CO2eq (g)", "PSNR", "SSIM", "VMAF")
        line = "+"
        for i in range(len(header1)-2):
            line = line+"-" 
        line = line+"+"
        line1 = list(line) 
        for i in [51, 62, 76, 91, 105, 114, 123]:
            line1[i] = "+" 
        print("")   
        ft = bg.green +'gr'+bg.rs
        print(fg.green+' '+ft+' : Encoded video is ready'+fg.rs)
        print("".join(line1))
        print(header1)
        print("".join(line1))
        
        for i in range(q[0], q[1]):
            video_or = sheet1.cell(row = i, column = 1).value
            video_name = sheet1.cell(row = i, column = 2).value
            bitratei = sheet1.cell(row = i, column = 3).value
            res = sheet1.cell(row = i, column = 4).value
            fps = sheet1.cell(row = i, column = 5).value
            pxl_fmt = sheet1.cell(row = i, column = 6).value
           
            interval = 0.3
            codec_list = codec_list_cpu
            psnr_col = 23
            if fps == 50 or fps == 60:
                bitrateIndex = [7, 11]
                #bitrateIndex = [10, 11]
            if fps == 20 or fps == 30:
                bitrateIndex = [2, 6]
                #bitrateIndex = [5, 6]
            #+' 2>&1 | grep "bench: utime" > '+self.path_envid+'/time.txt'
            pasCodec = 3
            input1 = self.path_envid+'/'+video_or
            #print(input1)
            if os.path.exists(input1) == True:
                for c in codec_list:
                    for b in range(bitrateIndex[0], bitrateIndex[1]):
                        ib = pasCodec + bitratei
                        #print(ib, b)
                        btr = sheet0.cell(row = ib, column = b).value
                        #print(btr)
                        os.makedirs(self.project_path+'/Encoded_video/', exist_ok=True)
                        os.makedirs(self.project_path+'/output/', exist_ok=True)
                        if c == 'VVenC' :  
                            vid_trans = video_name+'_'+str(btr)+'k_'+str(fps)+'fps_'+c+'.266'
                        else:
                            vid_trans = video_name+'_'+str(btr)+'k_'+str(fps)+'fps_'+c+'.mp4'
                        output1 = self.project_path+'/Encoded_video/'+vid_trans
                        #print(output1)
                        idv = self.find_index_row(sheet2, vid_trans)
                        otp = os.path.exists(output1)
                        flt = ''
                        if (otp == False or idv == None) :
                            if c == 'libx264' or c == 'libx265':
                                proc_name = 'ffmpeg'
                                if video_or.split('.')[-1] == 'yuv':
                                                              
                                    envid = 'ffmpeg -y -benchmark -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset veryslow -minrate '+str(btr)+'k -maxrate '+str(btr)+'k -b:v '+str(btr)+'k '+str(output1)+' 2>&1 | grep "bench: utime" > '+self.project_path+'/output/time.txt'
                                else :
                                    envid = 'ffmpeg -y -benchmark -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset veryslow -minrate '+str(btr)+'k -maxrate '+str(btr)+'k -b:v '+str(btr)+'k '+str(output1)+' 2>&1 | grep "bench: utime" > '+self.project_path+'/output/time.txt'
                         
                            elif c == 'libvpx-vp9':
                                proc_name = 'ffmpeg'
                                if video_or.split('.')[-1] == 'yuv':
                                    envid = 'ffmpeg  -y -benchmark -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -speed 0 -minrate '+str(btr)+'k -maxrate '+str(btr)+'k -b:v '+str(btr)+'k '+str(output1)+' 2>&1 | grep "bench: utime" > '+self.project_path+'/output/time.txt'
                                else :
                                    envid = 'ffmpeg -y -benchmark -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -speed 0 -minrate '+str(btr)+'k -maxrate '+str(btr)+'k -b:v '+str(btr)+'k '+str(output1)+' 2>&1 | grep "bench: utime" > '+self.project_path+'/output/time.txt'
                                    
                            
                            elif c == 'libaom-av1':
                                 proc_name = 'ffmpeg'
                                 if video_or.split('.')[-1] == 'yuv':
                                    envid = 'ffmpeg -y -benchmark -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -cpu-used 8 -minrate '+str(btr)+'k -maxrate '+str(btr)+'k -b:v '+str(btr)+'k -strict -2 '+str(output1)+' 2>&1 | grep "bench: utime" > '+self.project_path+'/output/time.txt'
                                 else :
                                    envid = 'ffmpeg -y -benchmark -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -cpu-used  8 -minrate '+str(btr)+'k -maxrate '+str(btr)+'k -b:v '+str(btr)+'k -strict -2 '+str(encod_path)+' 2>&1 | grep "bench: utime" > '+self.project_path+'/output/time.txt'
                            
                            elif c == 'libsvtav1':
                                 proc_name = 'ffmpeg'
                                 if video_or.split('.')[-1] == 'yuv':
                                    envid = 'ffmpeg -y -benchmark -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset 3 -b:v '+str(btr)+'k '+str(output1)+' 2>&1 | grep "bench: utime" > '+self.project_path+'/output/time.txt'
                                 else :
                                    envid = 'ffmpeg -y -benchmark -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset 3 -b:v '+str(btr)+'k '+str(output1)+' 2>&1 | grep "bench: utime" > '+self.project_path+'/output/time.txt'



                            elif c == 'VVenC' : 
                                proc_name = 'vvencapp'
                                if pxl_fmt == 'yuv420p10le' :
                                    pxl_fmt1 = 'yuv420_10'
                                    bitd = 10
                                elif pxl_fmt == 'yuv420p' :
                                    pxl_fmt1 = 'yuv420'
                                    bitd = 8
                                vid_trans = video_name+'_'+str(btr)+'k_'+str(fps)+'fps_'+c+'.266'
                                vid_trans_yuv = 'quality.yuv'
                                output1 = self.project_path+'/Encoded_video/'+vid_trans
                                output1_yuv = self.project_path+'/Encoded_video/'+vid_trans_yuv 
                                envid = VVenC_path[0]+' --preset medium -i '+str(input1)+' -c '+str(pxl_fmt1)+' -s '+str(res)+' --internal-bitdepth '+str(bitd)+' -r '+str(fps)+' -b '+str(btr)+'k -t 16 -o '+str(output1)+'  2>&1 > '+self.project_path+'/output/encVVC.txt'
                                devid = VVdeC_path[0]+' -b '+str(output1)+' -o '+str(output1_yuv)

                           
                            self.stop_threads = False
                            if os.path.exists('emissions.csv')== True:
                                os.remove('emissions.csv')
                            if os.path.exists('emissions.xlsx')== True:    
                                os.remove('emissions.xlsx')
                                        
                            
                            thread1 = threading.Thread(target = self.cpu_mem_usage_all, args=(proc_name, interval, ))
                            thread1.start()
                            t = time.time()
                            os.system("python3 /media/ridha/D81821F01821CE76/Doctorat/Projet4_Energy/Project_energy_final/CTC/code/energy.py -v '"+envid+"' > /dev/null 2>&1")
                            
                            tf = time.time()-t
                            self.stop_threads = True
                            thread1.join()
                            
                            if c == 'VVenC' : 
                                os.system(devid+" > /dev/null 2>&1")
                            
                            csv_file = 'emissions.csv'
                            xlsx_file = 'emissions.xlsx'
                            wb4 = Workbook()
                            ws = wb4.active
                            if os.path.exists(csv_file)== True:
                                if os.path.exists(xlsx_file)== False:
                                    with open(csv_file, 'r') as csv1:
                                        for row1 in csv.reader(csv1):
                                            ws.append(row1)
                                    wb4.save(xlsx_file)
                                if os.path.exists(xlsx_file)== True:
                                    wb5 = load_workbook(xlsx_file)
                                    sheet5 = wb5.worksheets[0]
                                else :
                                    print("xlsx file for CO2 emission not existe")
                            else :
                                print("csv file for CO2 emissionnot existe")
                            energy = sheet5.cell(row = 2, column = 6).value  
                            co2 = float(sheet5.cell(row = 2, column = 5).value)*1000
                            duration = sheet5.cell(row = 2, column = 4).value
                            if c == 'VVenC':
                                vbit = 14
                                btr = ''
                                while btr == '':
                                    bitrate = 'grep -i " a " '+self.project_path+'/output/encVVC.txt | cut -d " " -f'+str(vbit)+' > '+self.project_path+'/output/btr.txt'
                                    
                                    os.system(bitrate)
                                    a = open(self.project_path+'/output/btr.txt',"r")
                                    btr = str(a.readlines()[0].rstrip())
                                    vbit = vbit+1
                            
                                a.close()
                                t1 = '1=0'
                                t2 = '1=0'
                                timevvc = 'grep -i "Total Time" '+self.project_path+'/output/encVVC.txt | cut -d " " -f3 > '+self.project_path+'/output/time.txt'
                                os.system(timevvc)
                                a = open(self.project_path+'/output/time.txt',"r")
                                t3 = '1='+str(a.readlines()[0].rstrip()).replace("s", "")
                                a.close()
                            else:
                                bitrate = 'ffmpeg -i '+output1+' 2>&1 | grep "Duration" | cut -d " " -f8 > '+self.project_path+'/output/btr.txt'
                                os.system(bitrate)
                                # PSNR 
                                a = open(self.project_path+'/output/btr.txt',"r")
                                btr = str(a.readlines()[0].rstrip())
                                a.close()
                                a = open(self.project_path+'/output/time.txt',"r")
                                v2 = str(a.readlines()[0].rstrip()).replace("s", "")
                                a.close()
                                no, t1, t2, t3 = v2.split(' ')
                           
                            
                            
                           
                            if idv == None :
                                idv = sheet2.max_row + 1
                            sheet2.cell(row = idv, column = 1).value = vid_trans
                            sheet2.cell(row = idv, column = 2).value = float(t1.split('=')[-1]) #'Utime'
                            sheet2.cell(row = idv, column = 3).value = float(t2.split('=')[-1]) #'Time'
                            sheet2.cell(row = idv, column = 4).value = float(t3.split('=')[-1]) #'Rtime'
                            sheet2.cell(row = idv, column = 5).value = float(tf) #'Algo time'
                            sheet2.cell(row = idv, column = 6).value = int(float(btr)) #'Bitrate'
                            #print(self.info_sys_all)
                            info_cpu = np.array(self.info_sys_all, dtype="object")
                            min_info_cpu = self.average_column_matrix(info_cpu)
                            max_info_cpu = min_info_cpu
                            sheet2.cell(row = idv, column = 7).value = float(max_info_cpu[0]) #'%CPU system'
                            sheet2.cell(row = idv, column = 8).value = float(max_info_cpu[1]) #'%CPU FFmpeg'
                            sheet2.cell(row = idv, column = 9).value = int(max_info_cpu[2]) #'Total (MB)'
                            sheet2.cell(row = idv, column = 10).value = int(min_info_cpu[3]) #'Available (MB)'
                            sheet2.cell(row = idv, column = 11).value = int(min_info_cpu[4]) #'Free (MB)'
                            sheet2.cell(row = idv, column = 12).value = int(max_info_cpu[5]) #'Used system (MB)'
                            sheet2.cell(row = idv, column = 13).value = float(max_info_cpu[6]) #'%Used system'
                            sheet2.cell(row = idv, column = 14).value = int(max_info_cpu[7]) #'Used FFmpeg (MB)'
                            sheet2.cell(row = idv, column = 15).value = float(max_info_cpu[8]) #'%Used FFmpeg'
                            sheet2.cell(row = idv, column = 16).value = float(duration) #'Time
                            sheet2.cell(row = idv, column = 17).value = float(energy) #'Energy GPU'
                            sheet2.cell(row = idv, column = 18).value = float(co2) #'carbon'
                            sheet2.cell(row = idv, column = 19).value = float(co2)/122.4 #'This is equivalent to: km travelled by car'
                            psnr_col = 20
                            self.info_sys_all = []
                            self.info_sys_all_gpu = []
                            
                            tim = float(t3.split('=')[-1])
                            bit = int(float(btr))
                            enrg = float(energy)
                            co = float(co2)
                            vname = vid_trans
                            
                        else:
                            vname = fg.green +"{:^50}".format(vid_trans)+ fg.rs
                            tim = sheet2.cell(row = idv, column = 4).value
                            bit = sheet2.cell(row = idv, column = 6).value
                            enrg = sheet2.cell(row = idv, column = 17).value
                            co = sheet2.cell(row = idv, column = 18).value
                            psnr_col = 20
                           
                        vvmaf = ""
                        vpsnr = ""
                        vssim = ""
                     
                        if 'psnr' in metric:
                            vpsnr = sheet2.cell(row = idv, column = psnr_col).value
                            if otp == False or vpsnr == None:
                                if video_or.split('.')[-1] == 'yuv':
                                    if c == 'VVenC' : 
                                        command1 = 'ffmpeg -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(output1_yuv)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -filter_complex "'+flt+'psnr" -f null /dev/null 2>&1 | grep "PSNR y" | cut -d " " -f8 > '+self.project_path+'/output/f2.txt'
                                    else:
                                        command1 = 'ffmpeg -i '+str(output1)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -filter_complex "'+flt+'psnr" -f null /dev/null 2>&1 | grep "PSNR y" | cut -d " " -f8 > '+self.project_path+'/output/f2.txt'
                                    
       
                                else :
                                    command1 = 'ffmpeg -i '+str(output1)+' -i '+str(input1)+' -filter_complex "'+flt+'psnr" -f null /dev/null  2>&1 | grep "PSNR y" | cut -d " " -f8 > '+self.project_path+'/output/f2.txt'
                               
                                os.system(command1)
                                psnr = 'grep -i ":" '+self.project_path+'/output/f2.txt | cut -d ":" -f2 > '+self.project_path+'/output/psnr.txt'
                                os.system(psnr)
                                a = open(self.project_path+'/output/psnr.txt',"r")
                                vpsnr = float(a.readlines()[0].rstrip())
                                sheet2.cell(row = idv, column = psnr_col).value = float(vpsnr)
                            
                           
			            
                        if 'ssim' in metric:
                            vssim = sheet2.cell(row = idv, column = psnr_col+1).value
                            if otp == False or vssim == None:
                                if video_or.split('.')[-1] == 'yuv':
                                    if c == 'VVenC' : 
                                        command1 = 'ffmpeg -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(output1_yuv)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -filter_complex "'+flt+'ssim" -f null /dev/null 2>&1 | grep "SSIM Y" | cut -d " " -f11 > '+self.project_path+'/output/ssim1.txt'
                                    else:
                                        command1 = 'ffmpeg -i '+str(output1)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -filter_complex "'+flt+'ssim" -f null /dev/null 2>&1 | grep "SSIM Y" | cut -d " " -f11 > '+self.project_path+'/output/ssim1.txt'
                                else :
                                    command1 = 'ffmpeg -i '+str(output1)+' -i '+str(input1)+' -filter_complex "'+flt+'ssim" -f null /dev/null  2>&1 | grep "SSIM y" | cut -d " " -f8 > '+self.project_path+'/output/ssim1.txt'
                                os.system(command1)
                                ssim = 'grep -i ":" '+self.project_path+'/output/ssim1.txt | cut -d ":" -f2 > '+self.project_path+'/output/ssim.txt'
                                os.system(ssim)
                                a = open(self.project_path+'/output/ssim.txt',"r")
                                vssim = float(a.readlines()[0].rstrip())
                                sheet2.cell(row = idv, column = psnr_col+1).value = float(vssim)
                            
                          

                        if 'vmaf' in metric:
                            vvmaf = sheet2.cell(row = idv, column = psnr_col+2).value
                            if otp == False or vvmaf == None:
                                if bitratei == 1:
                                    model = 'ffmpeg/model/vmaf_4k_v0.6.1.json'
                                else : 
                                    model ='ffmpeg/model/vmaf_v0.6.1.json'
                                if video_or.split('.')[-1] == 'yuv':
                                    if c == 'VVenC' : 
                                        command1 = 'ffmpeg -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(output1_yuv)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -lavfi "'+flt+'libvmaf" -f null /dev/null 2>&1 | grep "VMAF score:" | cut -d ":" -f2 > '+self.project_path+'/output/vmaf1.txt'
                                    else:
                                        command1 = 'ffmpeg -i '+str(output1)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -lavfi "'+flt+'libvmaf" -f null /dev/null 2>&1 | grep "VMAF score:" | cut -d ":" -f2 > '+self.project_path+'/output/vmaf1.txt'
                                else :
                                    command1 = 'ffmpeg -i '+str(output1)+' -i '+str(input1)+' -lavfi "'+flt+'libvmaf" -f null /dev/null  2>&1 | grep "VMAF score:" | cut -d ":" -f2 > '+self.project_path+'/output/vmaf1.txt'
                                
                                os.system(command1)
                                a = open(self.project_path+'/output/vmaf1.txt',"r")
                                vvmaf = float(a.readlines()[0].rstrip())
                                sheet2.cell(row = idv, column = psnr_col+2).value = float(vvmaf)
                            
                        
                        core = "|{:^50}|{:^10.4}| {:^11} | {:^10.12} |{:^13.11}|{:^8.2f}|{:^8.2f}|{:^8.2f}|"
                        print(core.format(vname, tim, bit, str(enrg), str(co), vpsnr, vssim ,vvmaf))
                        print("".join(line1))
                        wb2.save(self.results)
			            
                    pasCodec = pasCodec + 8
            else:
                print('raw video not exist')
                    
#test.............................................................................................

ap = argparse.ArgumentParser()
ap.add_argument("-s", "--video_path", help="Path to dataset")
ap.add_argument("-r", "--results", help="Path to results")
#ap.add_argument("-e", "--results", help="Path to vvencapp")
#ap.add_argument("-d", "--results", help="Path to vvdecapp")
ap.add_argument('-m', '--metric', nargs='+', type=str, help='Metric quatlity')

args = vars(ap.parse_args())
q = [9, 10]
bitrate_target = os.path.dirname(args["video_path"])+"/bitrate-final.xlsx"
print(bitrate_target)
video_info = os.path.dirname(args["video_path"])+"/video_information.xlsx"

if os.path.exists(bitrate_target)== True:
    wb0 = load_workbook(bitrate_target)
else:
    print('XLSX file of target bitrate not exist')

if os.path.exists(video_info)== True:
    wb1 = load_workbook(video_info)
else:
    print('XLSX file of video information not exist')
    
if os.path.exists(args["results"])== True:
    wb2 = load_workbook(args["results"])
    sheet2 = wb2.worksheets[0]
    row_count = sheet2.max_row
    print(row_count) 
else:
    print('XLSX file of encoding results not exist')
    
sheet2.cell(row = 1, column = 1).value = 'Videos'
sheet2.cell(row = 1, column = 2).value = 'Utime (s)'
sheet2.cell(row = 1, column = 3).value = 'Time (s)'
sheet2.cell(row = 1, column = 4).value = 'Encoding time (s)'
sheet2.cell(row = 1, column = 5).value = 'Algo time (s)'   
sheet2.cell(row = 1, column = 6).value = 'Bitrate (kb/s)'
sheet2.cell(row = 1, column = 7).value = '%CPU system'
sheet2.cell(row = 1, column = 8).value = 'CPU utilization (%)'
sheet2.cell(row = 1, column = 9).value = 'Total (MB)'
sheet2.cell(row = 1, column = 10).value = 'Available (MB)'
sheet2.cell(row = 1, column = 11).value = 'Free (MB)'
sheet2.cell(row = 1, column = 12).value = 'Used system (MB)'
sheet2.cell(row = 1, column = 13).value = '%Memory system'
sheet2.cell(row = 1, column = 14).value = 'Used FFmpeg (MB)'
sheet2.cell(row = 1, column = 15).value = '%Memory utilization (%)'
sheet2.cell(row = 1, column = 16).value = 'Encoding time (s)' 
sheet2.cell(row = 1, column = 17).value = 'Energy (Kwh)'
sheet2.cell(row = 1, column = 18).value = 'CO2 emissions in France (g)'
sheet2.cell(row = 1, column = 19).value = 'Equivalent of CO2 emissions in distance travelled by car (m)'
mc = 0
if 'psnr' in args['metric']:
    sheet2.cell(row = 1, column = 20).value = 'PSNR'
    mc = 1
if 'ssim' in args['metric']:
    sheet2.cell(row = 1, column = 20+1).value = 'SSIM'
    mc = 2
if 'vmaf' in args['metric']:
    sheet2.cell(row = 1, column = 20+2).value = 'VMAF'
    mc = 3

    
wb2.save(args["results"]) 

p1 = video_encoding(args['results'], args['video_path'])
p1.encoding(args['metric'], q, wb0, wb1, wb2)



""" 
python3 /your-path-to-the-project/encoding_video.py
-s /your-path-to-the-project/original_dataset \
-r '/your-path-to-the-project/encoding_results.xlsx' \
-m psnr ssim vmaf

"""






