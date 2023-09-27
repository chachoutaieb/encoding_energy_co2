# @Developed by Chachou Taieb (°_^)
#!/usr/bin/env python3
import os
import sys
import argparse
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import psutil
import glob
import time
import subprocess
import numpy as np
import threading
from codecarbon import OfflineEmissionsTracker
import csv
from pynvml import *
from sty import fg, bg, ef, rs
np.set_printoptions(suppress=True)
#..............................................................................................
class video_encoding:
        
    def __init__(self, results, path_envid, project_path):
        self.results = results
        self.path_envid = path_envid
        self.project_path = project_path 
        self.info_sys = []
        self.info_sys_all = []
        self.info_sys_gpu = []
        self.info_sys_all_gpu = []
        self.stop_threads = False 
        
    def find_index_row(self, sheet, title_row):
        for i in range(1, sheet.max_row + 1):
            cell_name = sheet.cell(row = i, column = 1).value
            if title_row == cell_name:
                return i  
    
    def encoding_config(self, video_name, l1, param):
        
        
        f = open(video_name, "w")
        f.write("#======== File I/O ===============\n")
        for i, p in enumerate(l1):
            f.write(p+" : "+param[i]+"\n")
        f.close()
        
    # decorater used to block function printing to the console
    def blockPrinting(self, func):
        def func_wrapper(*args, **kwargs):
            # block all printing to the console
            sys.stdout = open(os.devnull, 'w')
            # call the method in question
            value = func(*args, **kwargs)
            # enable all printing to the console
            sys.stdout = sys.__stdout__
            # pass the return value of the method back
            return value
        return func_wrapper      
    
    def average_column_matrix(self, matrix):
        average = []
        for j in range(len(matrix[0])) :
            count = 0
            sumv = 0
            ave = 0
            for i in range(len(matrix)):
                if matrix[i][j] != 0.0 :
                    count = count+1
                    sumv = sumv + matrix[i][j]
            if count != 0:
                ave  = sumv / float(count)
            average.append(ave)
            #print('average(', matrix[:, j], ') =', ave)
        return average

    def get_pid(self, process_name):
        try:
            l = [int(i) for i in subprocess.check_output(["pidof",process_name]).split()] 
        except subprocess.CalledProcessError as e:
            l = []
            #print("The process does not exist")
        return l

    def cpu_mem_usage_all(self, proc_name, interval):
        while True:
            pids = self.get_pid(proc_name)
            if pids:
                try:
                    self.info_sys = []
                    proc = psutil.Process(pids[0])
                    #self.info_pid.append(float(pids[0]))
                    self.info_sys.append(float(psutil.cpu_percent(interval)))
                    self.info_sys.append(float(proc.cpu_percent(interval)))
                    self.info_sys.append(float(psutil.virtual_memory().total) / float(1024 ** 2)) 
                    self.info_sys.append(float(psutil.virtual_memory().available) / float(1024 ** 2)) 
                    self.info_sys.append(float(psutil.virtual_memory().free) / float(1024 ** 2)) 
                    self.info_sys.append(float(psutil.virtual_memory().used) / float(1024 ** 2))
                    self.info_sys.append(float(psutil.virtual_memory().percent))
                    self.info_sys.append(proc.memory_info().rss / float(1024 ** 2))
                    self.info_sys.append(float(proc.memory_percent()))
                    self.info_sys_all.append(self.info_sys)   
                except psutil.NoSuchProcess as e:
                    inor=1
                    #print(e)
            if self.stop_threads:
                break
    
    def gpu_cpu_info_all(self, interval):
        handle = nvmlDeviceGetHandleByIndex(0)
        while(True):
            pids_cpu = self.get_pid('ffmpeg')
            try:
                pids = nvmlDeviceGetComputeRunningProcesses(handle)
            except NVMLError as err:
                pids =  []    
            if pids:
                try:
                    self.info_sys = []
                    #self.info_pid.append(float(pids[0]))
                    self.info_sys.append(float(psutil.cpu_percent(interval)))
                    self.info_sys.append(float(psutil.virtual_memory().total) / float(1024 ** 2)) 
                    self.info_sys.append(float(psutil.virtual_memory().available) / float(1024 ** 2)) 
                    self.info_sys.append(float(psutil.virtual_memory().free) / float(1024 ** 2)) 
                    self.info_sys.append(float(psutil.virtual_memory().used) / float(1024 ** 2))
                    self.info_sys.append(float(psutil.virtual_memory().percent))
                    if pids_cpu:
                        proc = psutil.Process(pids_cpu[0])
                        self.info_sys.append(1.0)
                        self.info_sys.append(float(proc.cpu_percent(interval)))   
                        self.info_sys.append(proc.memory_info().rss / float(1024 ** 2))
                        self.info_sys.append(float(proc.memory_percent()))
                    else :
                        self.info_sys.append(0.0) 
                        self.info_sys.append(0.0)   
                        self.info_sys.append(0.0)
                        self.info_sys.append(0.0)
                    self.info_sys_all.append(self.info_sys)
                except psutil.NoSuchProcess as e:
                    inor=1
                    #print(e)
            if self.stop_threads:
                break

    def gpu_mem_usage(self, interval):
        handle = nvmlDeviceGetHandleByIndex(0)
        while(True):
            try:
                pids = nvmlDeviceGetComputeRunningProcesses(handle)
            except NVMLError as err:
                pids =  []    
            if pids:
                name = ''
                for p in pids:
                    try:
                        name = pynvml.nvmlSystemGetProcessName(p.pid).decode('utf-8')
                    except:
                        name = ''         
                time.sleep(interval)
                try:
                    temp = nvmlDeviceGetTemperature(handle, NVML_TEMPERATURE_GPU)
                    info = nvmlDeviceGetMemoryInfo(handle)
                    used = nvmlDeviceGetUtilizationRates(handle)
                    self.info_sys_gpu = []
                    self.info_sys_gpu.append(float(used.gpu))
                    self.info_sys_gpu.append(float(info.total / 1024 / 1024))
                    self.info_sys_gpu.append(float(info.free / 1024 / 1024))
                    self.info_sys_gpu.append(float(info.used / 1024 / 1024))
                    self.info_sys_gpu.append(float(pids[0].usedGpuMemory / 1024 / 1024))
                    self.info_sys_gpu.append(float(used.memory))
                    self.info_sys_gpu.append(float(temp))
                    self.info_sys_all_gpu.append(self.info_sys_gpu)
                except NVMLError as err:
                    temp = handleError(err)
            if self.stop_threads:
                break
	
#..............................................................................................	
    def encoding(self, metric, q, platform, wb0, wb1, wb2, pathV, pxl_fmtV, bitV, resV, fpsV, codec, speed, output1s, intrvalv):
        print("")
        if platform != 'GPU':
            
            VVenC_path = None
            VVdeC_path = None
            
            """
            Example :
                VVenC_path = 'path to folder/vvenc-master'
                VVdeC_path = 'path to folder/vvdec-master'
            """
            if VVenC_path == None or VVdeC_path ==  None:
                print(fg.red +'Error: Please install VVenC and ensure that the path to "vvenc-master" is correctly set on line 183 (encoding_energy_co2.py).'+fg.rs)
                print(fg.red +'Error: Please install VVdeC and ensure that the path to "vvdec-master" is correctly set on line 184 (encoding_energy_co2.py).\n'+fg.rs)
                exit()
        if pathV != None:
            sheet0 = ""
            sheet1 = ""
        else:
            sheet0 = wb0.worksheets[0]
            sheet1 = wb1.worksheets[0]
        sheet2 = wb2.worksheets[0]
        self.stop_threads = False
        proc_name = 'ffmpeg'
        #'libx265', 
        codec_name = ['x264', 'x265', 'vp9', 'vvenc', 'svt-av1']#, 
        codec_cpu = ['libx264' ,'libx265', 'libvpx-vp9', 'VVenC', 'libsvtav1'] 
        codec_gpu = ['h264_nvenc', 'hevc_nvenc']
        codec_list_cpu = []
        codec_list_gpu = []
        gopsize = [32, 32, 32, 64, 64, 96]
        fpslist = [20, 24, 30, 50, 60, 100]
        if  platform == 'CPU':
            if speed == 'slower':
                confpath = VVenC_path+'/cfg/randomaccess_slow.cfg'
                confpath64 = self.project_path+'/data/randomaccess_slow64.cfg'
                preset = ['veryslow', '0', '3', 'slow']
                interval_codec = [0.3, 0.3, 2, 3, 0.3]
            elif speed == 'medium':
                confpath = VVenC_path+'/cfg/randomaccess_medium.cfg'
                confpath64 = self.project_path+'/data/randomaccess_medium64.cfg'
                preset = ['medium', '3', '6', 'medium']
                interval_codec = [0.06,  0.15, 0.5, 3, 0.5]
            elif speed == 'fast':
                confpath = VVenC_path+'/cfg/randomaccess_fast.cfg'
                confpath64 =  self.project_path+'/data/randomaccess_fast64.cfg'
                preset = ['veryfast', '6', '10', 'fast']
                interval_codec = [0.04,  0.1, 0.5, 1, 0.2]
            elif speed == 'faster':
                confpath = VVenC_path+'/cfg/randomaccess_faster.cfg'
                confpath64 =  self.project_path+'/data/randomaccess_faster64.cfg'
                preset = ['ultrafast', '8', '12', 'faster']
                interval_codec = [0.04,  0.1, 0.2, 1, 0.2]
            if intrvalv != None:
                if pathV == None:
                    if len(intrvalv) == 5:
                        interval_codec = intrvalv
                    else:
                        print(fg.red +"Error: please provide five values for the five codecs (x264, x265, VP9, NVENC, and SVT-AV1)  ex. -it 0.3 0.3 2 3 0.5.\n"+fg.rs)
                        exit()
                else:
                    for v in range(4):
                        intrvalv.insert(0, intrvalv[0])
                    interval_codec = intrvalv
            if  speed not in ['slower', 'medium' ,'fast', 'faster']:
                print(fg.red +'Error: please select one preset from this list {}.\n'.format(['slower', 'medium' ,'fast', 'faster'])+fg.rs )
                exit() 
        elif platform == 'GPU':
            if speed == 'slow':
                interval_codec = [0.03, 0.05]
            elif speed == 'medium':
                interval_codec = [0.03,  0.05]
            elif speed == 'fast':
                interval_codec = [0.03,  0.05]
            if intrvalv != None:
                if pathV == None:
                    if len(intrvalv) == 2:
                        interval_codec = intrvalv
                    else:
                        print(fg.red +"Error: please provide two values for the two codecs (h264_nvenc and hevc_nvenc)  ex. -it 0.03 0.05.\n"+fg.rs)
                        exit()
                else:
                    for v in range(1):
                        intrvalv.insert(0, intrvalv[0])
                    interval_codec = intrvalv
            preset = [speed]
            if  speed not in ['slow', 'medium' ,'fast']:
                print(fg.red +'Error: please select one preset from this list {}.\n'.format(['slow', 'medium' ,'fast'])+fg.rs )
                exit()
        if pathV != None:
            exi = False
            if output1s == None:
                print(fg.red +"Error: please provide the path to the outputideo.mp4.\n"+fg.rs)
                exi = True
            if pxl_fmtV == None:
                print(fg.red +"Error: please provide the bit depth  of original video (8 or 10).\n"+fg.rs)
                exi = True
            if bitV == None:
                print(fg.red +"Error: please provide the encoding bitrate (kb/s.\n"+fg.rs)
                exi = True
            if resV == None:
                print(fg.red +"Error: please provide the resolution  of original video WxH).\n"+fg.rs)
                exi = True
            if fpsV == None:
                print(fg.red +"Error: please provide the framerate of original video.\n"+fg.rs)
                exi = True
            if codec == None:
                print(fg.red +"Error: please provide the  type of codec.\n"+fg.rs)
                exi = True    
            if speed == None:
                print(fg.red +"Error: please provide the  encoding preset.\n"+fg.rs)
                exi = True        
            if exi == True:
                exit()            
            if  platform == 'CPU':
                if  speed not in ['slower', 'medium' ,'fast', 'faster']:
                    print(fg.red +'Error: please select one preset from this list {}.\n'.format(['slower', 'medium' ,'fast', 'faster'])+fg.rs )
                    exit()
                if codec in codec_name:
                    ind = codec_name.index(codec)
                    codec_list_cpu = [codec_cpu[ind]]
                else :
                    print(fg.red +'Error: please select one codec from this list {}.\n'.format(codec_name)+fg.rs)
                    exit()
            else:
                if  speed not in ['slow', 'medium' ,'fast']:
                    print(fg.red +'Error: please select one preset from this list {}.\n'.format(['slow', 'medium' ,'fast'])+fg.rs )
                    exit()
                if codec in codec_gpu:
                    ind = codec_gpu.index(codec)
                    codec_list_gpu = [codec_gpu[ind]]
                else :
                    print(fg.red +'Error: please select one codec from this list {}.\n'.format(codec_gpu[0:2])+fg.rs)
                    exit()
            print('\nThe energy value is calculated with the following interval period: {} = {}s.\n You can change it according to the machine\'s performance by using the "-it" option.\n'.format(codec, interval_codec[0]))
            print("") 
        else :
            if  platform == 'CPU':
                print('\nThe energy value is calculated with the following interval period: x264 = {}s, x265 = {}s, vp9 = {}s, VVenC = {}s, SVT-AV1 = {}s.\n You can change them according to the machine\'s performance by using the "-it" option.\n'.format(interval_codec[0], interval_codec[1], interval_codec[2], interval_codec[3], interval_codec[4]))
                print("")   
            else:    
                print('\nThe energy value is calculated with the following interval period: h264_nvenc = {}s, hevc_nvenc = {}s.\n You can change them according to the machine\'s performance by using the "-it" option.\n'.format(interval_codec[0], interval_codec[1]))
                print("") 
            codec_list_cpu = ['libx264', 'libx264', 'libx265','libvpx-vp9', 'VVenC', 'libsvtav1']#, 'libx264', 'libx265', , 'VVenC', 'libsvtav1'
            codec_list_gpu = ['h264_nvenc', 'hevc_nvenc']
        header1 = "|{:^60}|{:^10}|{:^13}|{:^14}|{:^13}|{:^8}|{:^8}|{:^8}|".format("Encoded video", "Time (s)", "Bitrate(kb/s)", "Energy (Wh)", "CO2eq (g)", "PSNR", "SSIM", "VMAF")
        line = "+"
        for i in range(len(header1)-2):
            line = line+"-" 
        line = line+"+"
        line1 = list(line) 
        for i in [61, 72, 86, 101, 115, 124, 133]:
            line1[i] = "+" 
        ft = bg.red +'gr'+bg.rs
        print(fg.red+' '+ft+' : Raw video does not exist.'+fg.rs)
        ft = bg.green +'gr'+bg.rs
        print(fg.green+' '+ft+' : Encoded video is ready.'+fg.rs)
        print("".join(line1))
        print(header1)
        print("".join(line1))
        qplist1 = [22, 27, 32, 37]
        qplist2 = [33, 42, 51, 59]
        #lq = 
        #q = list(range(3,11))+list(range(14,16))+list(range(20,22))
        for i in range(2, 28):
            if pathV != None:
                video_or = pathV
                video_name = os.path.basename(os.path.normpath(pathV)).split('.')[0]
                bitratei = bitV
                res = resV
                fps = fpsV
                pxl_fmt = pxl_fmtV
                nbr_frame = -1
            else:
                video_or = sheet1.cell(row = i, column = 1).value
                video_name = sheet1.cell(row = i, column = 2).value
                bitratei = sheet1.cell(row = i, column = 3).value
                res = sheet1.cell(row = i, column = 4).value
                fps = sheet1.cell(row = i, column = 5).value
                pxl_fmt = sheet1.cell(row = i, column = 6).value
                nbr_frame = sheet1.cell(row = i, column = 7).value
            if platform == 'GPU':
                codec_list = codec_list_gpu
                psnr_col = 12
            elif platform == 'CPU':
                codec_list = codec_list_cpu
                psnr_col = 9
            if fps == 50 or fps == 60:
                bitrateIndex = [7, 11]
                #bitrateIndex = [10, 11]
            if fps == 20 or fps == 30:
                bitrateIndex = [2, 6]
                #bitrateIndex = [5, 6]
            if pathV != None:
                bitrateIndex = [2, 3]
            #+' 2>&1 | grep "bench: utime" > '+self.path_envid+'/time.txt'
            pasCodec = 3
            if pathV != None:
                input1 = video_or
            else:
                input1 = os.path.join(self.path_envid, video_or)
            #print(input1)
            if os.path.exists(input1) == True:
                for ico, c in enumerate(codec_list):
                    if c in ['libx264' ,'libx265', 'VVenC']:
                        qplist = qplist1
                    else:
                        qplist = qplist2
                    
                    for iqp, qp in enumerate(qplist):
                        """
                        ib = pasCodec + bitratei
                        #print(ib, b)
                        if pathV != None:
                            btr = bitV
                        else :  
                            btr = sheet0.cell(row = ib, column = b).value
                        #print(btr)
                        """
                        os.makedirs(os.path.join(self.project_path, 'encoded_video_'+speed+'2'), exist_ok=True)
                        os.makedirs(os.path.join(self.project_path, 'output'), exist_ok=True)
                        if c == 'VVenC' :  
                            vid_trans = video_name+'_QP'+str(qp)+'_'+str(fps)+'fps_'+c+'_'+speed+'.266'
                        elif ico == 0 and iqp == 0:
                            vid_trans = video_name+'_QP'+str(qp)+'_'+str(fps)+'fps_'+c+'1_'+speed+'.mp4'
                        else:
                            vid_trans = video_name+'_QP'+str(qp)+'_'+str(fps)+'fps_'+c+'_'+speed+'.mp4'
                        if pathV != None:
                            output1 = output1s.split('.')[0]+'_QP'+str(qp)+'_'+c+'_'+speed+'.'+output1s.split('.')[1]
                            vid_trans = os.path.basename(output1s).split('.')[0]+'_QP'+str(qp)+'_'+c+'_'+speed+'.'+os.path.basename(output1s).split('.')[1]
                        else:
                            output1 = os.path.join(self.project_path, 'encoded_video_'+speed+'2', vid_trans)
                        idv = self.find_index_row(sheet2, vid_trans)
                        otp = os.path.exists(output1)
                        flt = ''
                        encVVCtxt = os.path.join(self.project_path, 'output', vid_trans+'.txt')
                        if idv == None :
                            time_p = os.path.join(self.project_path, 'output', 'time.txt')
                            idfps = fpslist.index(fps)
                            gs = 32
                            if c =='libx265':
                                interval = interval_codec[1]
                                proc_name = 'ffmpeg'
                                if video_or.split('.')[-1] == 'yuv':
                                    envid =  'ffmpeg -y -benchmark -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset '+preset[0]+' -frames:v '+str(nbr_frame)+' -x265-params "keyint='+str(gs)+':min-keyint='+str(gs)+'" -crf '+str(qp)+' '+str(output1)+' 2>&1 | grep "bench: utime" > '+time_p    
                                else :
                                    envid =  'ffmpeg -y -benchmark -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset '+preset[0]+' -frames:v '+str(nbr_frame)+' -x265-params "keyint='+str(gs)+':min-keyint='+str(gs)+'" -qp '+str(qp)+' '+str(output1)+' 2>&1 | grep "bench: utime" > '+time_p
                                #print(envid)
                            elif c =='libx264':
                                interval = interval_codec[0]
                                proc_name = 'ffmpeg'
                                if video_or.split('.')[-1] == 'yuv':
                                    envid =  'ffmpeg -y -benchmark -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset '+preset[0]+' -frames:v '+str(nbr_frame)+' -x264-params  "keyint='+str(gs)+':min-keyint='+str(gs)+'" -crf '+str(qp)+' '+str(output1)+' 2>&1 | grep "bench: utime" > '+time_p    
                                else :
                                    envid =  'ffmpeg -y -benchmark -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset '+preset[0]+' -frames:v '+str(nbr_frame)+' -x264-params  "keyint='+str(gs)+':min-keyint='+str(gs)+'" -qp '+str(qp)+' '+str(output1)+' 2>&1 | grep "bench: utime" > '+time_p
                            elif c == 'libvpx-vp9':
                                proc_name = 'ffmpeg'
                                interval = interval_codec[2]
                                if speed in ['slower', 'medium', 'fast']:
                                    if video_or.split('.')[-1] == 'yuv':
                                        envid =  'ffmpeg -y -benchmark -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -speed '+preset[1]+' -frames:v '+str(nbr_frame)+' -b:v 0 -crf '+str(qp)+' -g '+str(gs)+' -keyint_min '+str(gs)+' '+str(output1)+' 2>&1 | grep "bench: utime" > '+time_p
                                    else :
                                        envid =  'ffmpeg -y -benchmark -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -speed '+preset[1]+' -frames:v '+str(nbr_frame)+' -b:v 0 -crf '+str(qp)+' -g '+str(gs)+' -keyint_min '+str(gs)+' '+str(output1)+' 2>&1 | grep "bench: utime" > '+time_p      
                                elif speed in ['faster']:
                                    if video_or.split('.')[-1] == 'yuv':
                                        envid =  'ffmpeg  -y -benchmark -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -quality realtime -speed '+preset[1]+' -frames:v '+str(nbr_frame)+' -b:v 0 -crf '+str(qp)+' -g '+str(gs)+' -keyint_min '+str(gs)+' '+str(output1)+' 2>&1 | grep "bench: utime" > '+time_p
                                    else :
                                        envid =  'ffmpeg -y -benchmark -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -quality realtime -speed '+preset[1]+' -frames:v '+str(nbr_frame)+' -b:v 0 -crf '+str(qp)+' -g '+str(gs)+' -keyint_min '+str(gs)+' '+str(output1)+' 2>&1 | grep "bench: utime" > '+time_p    
                            
                            elif c == 'libsvtav1':
                                interval = interval_codec[4]
                                proc_name = 'ffmpeg'
                                if video_or.split('.')[-1] == 'yuv':
                                    envid =  'ffmpeg -y -benchmark -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset '+preset[2]+' -frames:v '+str(nbr_frame)+' -crf '+str(qp)+' -g '+str(gs)+' '+str(output1)+' 2>&1 | grep "bench: utime" > '+time_p
                                else :
                                    envid =  'ffmpeg -y -benchmark -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset '+preset[2]+' -frames:v '+str(nbr_frame)+' -crf '+str(qp)+' -g '+str(gs)+' '+str(output1)+' 2>&1 | grep "bench: utime" > '+time_p
                            elif c == 'VVenC' : 
                                interval = interval_codec[3]
                                proc_name = 'vvencFFapp'
                                if pxl_fmt == 'yuv420p10le' :
                                    pxl_fmt1 = 'yuv420_10'
                                    bitd = 10
                                elif pxl_fmt == 'yuv420p' :
                                    pxl_fmt1 = 'yuv420'
                                    bitd = 8                                
                                vid_trans_yuv = 'quality.yuv'
                                output1 = os.path.join(self.project_path, 'encoded_video_'+speed+'2', vid_trans)
                                output1_yuv = os.path.join(self.project_path, vid_trans_yuv) 
                                w, h = res.split('x')
                                l1 = ['InputFile', 'InputBitDepth', 'InputChromaFormat', 'FrameRate', 'FrameScale', 'FrameSkip', 'SourceWidth', 'SourceHeight', 'QP', 'Threads', 'FramesToBeEncoded',  'BitstreamFile']
                                param = [input1, str(bitd), '420',  str(fps), '1', '0', w, h, str(qp), '16', str(nbr_frame), output1]
                                squenceconfig = os.path.join(self.project_path, 'output', 'sequence.cfg')
                                self.encoding_config(squenceconfig, l1, param)
                                if fps >= 50:
                                    envid = VVenC_path+'/bin/release-static/vvencFFapp'+' -c '+confpath64+' -c '+squenceconfig+' 2>&1 > '+encVVCtxt 
                                else:
                                    envid = VVenC_path+'/bin/release-static/vvencFFapp'+' -c '+confpath+' -c '+squenceconfig+' 2>&1 > '+encVVCtxt 
                                 
                                devid = VVdeC_path+'/bin/release-static/vvdecapp'+' -b '+str(output1)+' -o '+str(output1_yuv) 
                                #print(envid) 
                            """               
                            elif c == 'hevc_nvenc' :
                                interval = interval_codec[1]
                                if video_or.split('.')[-1] == 'yuv':
                                    envid = 'ffmpeg -y -benchmark  -vsync 0 -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset '+preset[0]+' -minrate '+str(btr)+'k -maxrate '+str(btr)+'k -b:v '+str(btr)+'k '+output1+' 2>&1 | grep "bench: utime" > '+time_p
                                else:
                                    envid = 'ffmpeg -y -benchmark -hwaccel auto  -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str(pxl_fmt)+' -preset '+preset[0]+' -minrate '+str(btr)+'k -maxrate '+str(btr)+'k -b:v '+str(btr)+'k '+output1+' 2>&1 | grep "bench: utime" > '+time_p 
                            elif c == 'h264_nvenc' :
                                interval = interval_codec[0]
                                if video_or.split('.')[-1] == 'yuv':
                                    envid = 'ffmpeg -y -benchmark  -vsync 0 -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -c:v '+c+'  -pix_fmt '+str('yuv420p')+' -preset '+preset[0]+' -minrate '+str(btr)+'k -maxrate '+str(btr)+'k -b:v '+str(btr)+'k '+output1+' 2>&1 | grep "bench: utime" > '+time_p
                                else:
                                    envid = 'ffmpeg -y -benchmark -hwaccel auto  -i '+str(input1)+' -c:v '+c+' -pix_fmt '+str('yuv420p')+' -preset '+preset[0]+' -minrate '+str(btr)+'k -maxrate '+str(btr)+'k -b:v '+str(btr)+'k '+output1+' 2>&1 | grep "bench: utime" > '+time_p
                                #flt = '[1:v]format=pix_fmts=yuv420p[ref]; [ref]'
                            """
                            
                            self.stop_threads = False
                            if os.path.exists('emissions.csv')== True:
                                os.remove('emissions.csv')
                            if os.path.exists('emissions.xlsx')== True:    
                                os.remove('emissions.xlsx')         
                            if platform== 'CPU':
                                thread1 = threading.Thread(target = self.cpu_mem_usage_all, args=(proc_name, interval, ))
                                thread1.start()
                            elif platform == 'GPU':
                                nvmlInit()
                                thread2 = threading.Thread(target = self.gpu_mem_usage, args=(interval, ))
                                thread3 = threading.Thread(target = self.gpu_cpu_info_all, args=(interval, ))
                                thread2.start()
                                thread3.start()
                            t = time.time()
                            
                            os.system("python3 "+os.path.join(self.project_path, 'energy.py')+" -s "+str(interval)+" -c 'FRA' -v '"+envid+"' > /dev/null 2>&1")
                            tf = time.time()-t
                            self.stop_threads = True
                            if platform == 'CPU':
                                thread1.join()
                            elif platform == 'GPU':
                                thread2.join()
                                thread3.join()
                            if c == 'VVenC' : 
                                os.system(devid+" > /dev/null 2>&1")
                            csv_file = 'emissions.csv'
                            xlsx_file = 'emissions.xlsx'
                            wb4 = Workbook()
                            ws = wb4.active
                            if os.path.exists(csv_file)== True:
                                if os.path.exists(xlsx_file)== False:
                                    with open(csv_file, 'r') as csv1:
                                        for row1 in csv.reader(csv1):
                                            ws.append(row1)
                                    wb4.save(xlsx_file)
                                if os.path.exists(xlsx_file)== True:
                                    wb5 = load_workbook(xlsx_file)
                                    sheet5 = wb5.worksheets[0]
                                else :
                                    print("xlsx file for CO2 emission not existe")
                            else :
                                print("csv file for CO2 emission not existe")
                            energy1 = sheet5.cell(row = 2, column = 13).value  #kwh
                            energy = float(energy1)*1000 #kwh to wh
                            co2 = (float(energy1)*101)# kg to g
                            duration = sheet5.cell(row = 2, column = 4).value
                            btrtxt = os.path.join(self.project_path, 'output', 'btr.txt')
                            btrtxt2 = os.path.join(self.project_path, 'output', 'btr2.txt')
                            if c == 'VVenC':
                                vbit = 15
                                btr = ''
                                while btr == '':
                                    bitrate = 'grep -i " a " '+encVVCtxt+' | cut -d "." -f1 > '+btrtxt
                                    
                                    os.system(bitrate)
                                    os.system('grep -i "a" '+btrtxt+' | cut -d "a" -f2 > '+btrtxt2)
                                    a = open(btrtxt2,"r")
                                    btr = str(a.readlines()[0].rstrip())
                                    vbit = vbit+1
                                a.close()
                                t1 = '1=0'
                                t2 = '1=0'
                                timevvc = 'grep -i "Total Time" '+encVVCtxt+' | cut -d "]" -f3 > '+time_p
                                os.system(timevvc)
                                a = open(time_p, "r")
                                tk = str(a.readlines()[0].rstrip()).replace("[elapsed", "")
                                tk = tk.replace("sec.", "")
                                t3 = '1='+tk
                                a.close()
                            else:
                                bitrate =  'ffmpeg -i '+output1+' 2>&1 | grep "Duration" | cut -d " " -f8 > '+btrtxt
                                os.system(bitrate)
                                # PSNR 
                                a = open(btrtxt, "r")
                                btr = str(a.readlines()[0].rstrip())
                                a.close()
                                a = open(time_p, "r")
                                v2 = str(a.readlines()[0].rstrip()).replace("s", "")
                                a.close()
                                no, t1, t2, t3 = v2.split(' ')
                            if idv == None :
                                idv = sheet2.max_row + 1
                            sheet2.cell(row = idv, column = 1).value = vid_trans
                            sheet2.cell(row = idv, column = 2).value = float(t3.split('=')[-1]) #'Rtime'
                            sheet2.cell(row = idv, column = 3).value = int(float(btr)) #'Bitrate'
                            if  platform == 'CPU':
                                #print(self.info_sys_all)
                                info_cpu = np.array(self.info_sys_all, dtype="object")
                                
                                try:
                                    min_info_cpu = self.average_column_matrix(info_cpu)
                                except:
                                    print(fg.red +"\nError: The energy interval period {}s is too large. Please reduce the value such that at a minimum, 2 * interval is less than the encoding time {}s.\n".format(interval, float(t3.split('=')[-1]))+fg.rs ) 
                                    exit()
                                max_info_cpu = min_info_cpu
                                sheet2.cell(row = idv, column = 4).value = float(max_info_cpu[1]) #'%CPU FFmpeg'
                                sheet2.cell(row = idv, column = 5).value = float(max_info_cpu[8]) #'%Used FFmpeg'
                                sheet2.cell(row = idv, column = 6).value = float(energy) #'Energy GPU'
                                sheet2.cell(row = idv, column = 7).value = float(co2) #'carbon'
                                sheet2.cell(row = idv, column = 8).value = (float(co2)/120.4)*1000 #'This is equivalent to: km travelled by car'
                                psnr_col = 9
                            elif platform == 'GPU':
                                info_gpu = np.array(self.info_sys_all_gpu, dtype="object")
                                try:
                                    min_info_gpu = self.average_column_matrix(info_gpu) #np.amin(info_gpu, axis=0)
                                except:
                                    print(fg.red +"\nError: The energy interval period {}s is too large. Please reduce the value such that at a minimum, 2 * interval is less than the encoding time {}s.\n".format(interval, float(t3.split('=')[-1]))+fg.rs ) 
                                    exit()
                                max_info_gpu = min_info_gpu # np.amax(info_gpu, axis=0)
                                info_cpu = np.array(self.info_sys_all)
                                try:
                                    min_info_cpu = self.average_column_matrix(info_cpu) #np.amin(info_cpu, axis=0)
                                except:
                                    print(fg.red +"\nError: The energy interval period {}s is too large. Please reduce the value such that at a minimum, 2 * interval is less than the encoding time {}s.\n".format(interval, float(t3.split('=')[-1]))+fg.rs ) 
                                    exit()
                                max_info_cpu = min_info_cpu #np.amax(info_cpu, axis=0)
                                sheet2.cell(row = idv, column = 4).value = float(max_info_cpu[7]) #'CPU utilization FFmpeg (%)'
                                sheet2.cell(row = idv, column = 5).value = float(max_info_cpu[9]) #'%Memory FFmpeg'
                                sheet2.cell(row = idv, column = 6).value = float(max_info_gpu[0]) #'%GPU'
                                sheet2.cell(row = idv, column = 7).value = float(max_info_gpu[5]) #'%Memory GPU'
                                sheet2.cell(row = idv, column = 8).value = int(max_info_gpu[6]) #'Temperature'
                                sheet2.cell(row = idv, column = 9).value = float(energy) #'Energy GPU'
                                sheet2.cell(row = idv, column = 10).value = float(co2) #'carbon'
                                sheet2.cell(row = idv, column = 11).value = (float(co2)/120.4)*1000 #'This is equivalent to: m travelled by car'
                                psnr_col = 12
                            self.info_sys_all = []
                            self.info_sys_all_gpu = []
                            tim = float(t3.split('=')[-1])
                            bit = int(float(btr))
                            enrg = float(energy)
                            co = float(co2)
                            vname = vid_trans
                        else:
                            if platform == 'CPU':
                                vname = fg.green +"{:^60}".format(vid_trans)+ fg.rs
                                tim = sheet2.cell(row = idv, column = 2).value
                                bit = sheet2.cell(row = idv, column = 3).value
                                enrg = sheet2.cell(row = idv, column = 6).value
                                co = sheet2.cell(row = idv, column = 7).value
                                psnr_col = 9
                            elif platform == 'GPU':
                                vname = fg.green +"{:^60}".format(vid_trans)+ fg.rs
                                tim = sheet2.cell(row = idv, column = 2).value
                                bit = sheet2.cell(row = idv, column = 3).value
                                enrg = sheet2.cell(row = idv, column = 9).value
                                co = sheet2.cell(row = idv, column = 10).value
                                psnr_col = 12
                        vvmaf = 0
                        vpsnr = 0
                        vssim = 0
                        f2txt = os.path.join(self.project_path, 'output', 'f2.txt')
                        
                        tgd = '/content/ffmpeg-6.0-amd64-static/'
                        
                        
                        try:
                            if 'psnr' in metric:
                                vpsnr = sheet2.cell(row = idv, column = psnr_col+3).value
                                if vpsnr == None or vpsnr == 0:
                                    if video_or.split('.')[-1] == 'yuv':
                                        command1 =  'ffmpeg -i '+str(output1)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -frames:v '+str(nbr_frame)+' -filter_complex "'+flt+'psnr" -f null /dev/null 2>&1 | grep "PSNR y"  > '+f2txt
                                    else :
                                        command1 =  'ffmpeg -i '+str(output1)+' -i '+str(input1)+' -filter_complex "'+flt+'psnr" -f null /dev/null  2>&1 | grep "PSNR y" > '+f2txt
                                    #print(command1)
                                    if c == 'VVenC':
                                        output266 =  os.path.join(self.project_path, 'output', vid_trans+'.txt')
                                        os.system('grep  -i " a " '+output266+' > psnrtxt.txt')
                                        a = open('psnrtxt.txt', "r")
                                        linep= a.readlines()[0].rstrip()
                                        with open('psnrtxt.txt', 'w') as f:
                                            f.write(" ".join(linep.split()))
                                        os.system('grep  -i " a " psnrtxt.txt | cut -d " " -f6  > psnrYtxt.txt')
                                        os.system('grep  -i " a " psnrtxt.txt | cut -d " " -f7  > psnrUtxt.txt')
                                        os.system('grep  -i " a " psnrtxt.txt | cut -d " " -f8  > psnrVtxt.txt')
                                        #print('grep  -i " a " psnrtxt.txt | cut -d " " -f18  > psnrYtxt.txt')
                                    else:
                                        os.system(command1)
                                        os.system('grep  -i "PSNR y" '+f2txt+' | cut -d " " -f5 | cut -d ":" -f2 > psnrYtxt.txt')
                                        os.system('grep  -i "PSNR y" '+f2txt+' | cut -d " " -f6 | cut -d ":" -f2 > psnrUtxt.txt')
                                        os.system('grep  -i "PSNR y" '+f2txt+' | cut -d " " -f7 | cut -d ":" -f2 > psnrVtxt.txt')
                                        
                                    a = open('psnrYtxt.txt', "r")
                                    psnrY= float(a.readlines()[0].rstrip())
                                    a = open('psnrUtxt.txt', "r")
                                    psnrU= float(a.readlines()[0].rstrip())
                                    a = open('psnrVtxt.txt', "r")
                                    psnrV= float(a.readlines()[0].rstrip())
                                    sheet2.cell(row = idv, column = psnr_col).value = float(psnrY)  
                                    sheet2.cell(row = idv, column = psnr_col+1).value = float(psnrU) 
                                    sheet2.cell(row = idv, column = psnr_col+2).value = float(psnrV) 
                                    vpsnr = (float(psnrY)*6+float(psnrU)+float(psnrV))/8
                                    sheet2.cell(row = idv, column = psnr_col+3).value = vpsnr
                            
                            if 'ssim' in metric:
                                vssim = sheet2.cell(row = idv, column = psnr_col+7).value
                                if vssim == None or vssim == 0:
                                    if video_or.split('.')[-1] == 'yuv':
                                        if c == 'VVenC' : 
                                            #print(devid)
                                            os.system(devid+" > /dev/null 2>&1")
                                            if pxl_fmt == 'yuv420p':
                                                command1 =  'ffmpeg -f rawvideo -pix_fmt yuv420p10le -s:v '+str(res)+' -r '+str(fps)+' -i '+str(output1_yuv)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -frames:v '+str(nbr_frame)+' -filter_complex "'+flt+'ssim" -f null /dev/null 2>&1 | grep "SSIM Y"  > '+f2txt
                                            else:
                                                command1 =  'ffmpeg -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(output1_yuv)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -frames:v '+str(nbr_frame)+' -filter_complex "'+flt+'ssim" -f null /dev/null 2>&1 | grep "SSIM Y"  > '+f2txt
                                            
                                        else:
                                            command1 =  'ffmpeg -i '+str(output1)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -frames:v '+str(nbr_frame)+' -filter_complex "'+flt+'ssim" -f null /dev/null 2>&1 | grep "SSIM Y"  > '+f2txt
                                    else :
                                        command1 =  'ffmpeg -i '+str(output1)+' -i '+str(input1)+' -filter_complex "'+flt+'ssim" -f null /dev/null  2>&1 | grep "SSIM Y"  > '+f2txt
                                    #print(command1)
                                    os.system(command1)
                                    
                                    os.system('grep  -i "SSIM Y" '+f2txt+' | cut -d " " -f5 | cut -d ":" -f2 > ssimYtxt.txt')
                                    os.system('grep  -i "SSIM Y" '+f2txt+' | cut -d " " -f7 | cut -d ":" -f2 > ssimUtxt.txt')
                                    os.system('grep  -i "SSIM Y" '+f2txt+' | cut -d " " -f9 | cut -d ":" -f2 > ssimVtxt.txt')
                                    
                                    a = open('ssimYtxt.txt', "r")
                                    ssimY= float(a.readlines()[0].rstrip())
                                    a = open('ssimUtxt.txt', "r")
                                    ssimU= float(a.readlines()[0].rstrip())
                                    a = open('ssimVtxt.txt', "r")
                                    ssimV= float(a.readlines()[0].rstrip())
                                    sheet2.cell(row = idv, column = psnr_col+4).value = float(ssimY)  
                                    sheet2.cell(row = idv, column = psnr_col+5).value = float(ssimU) 
                                    sheet2.cell(row = idv, column = psnr_col+6).value = float(ssimV) 
                                    vssim = (float(ssimY)*6+float(ssimU)+float(ssimV))/8
                                    sheet2.cell(row = idv, column = psnr_col+7).value = vssim 

                            if 'vmaf' in metric:
                                vvmaf = sheet2.cell(row = idv, column = psnr_col+8).value
                                if  vvmaf == None or  vvmaf == 0:
                                    if bitratei == 1:
                                        model = 'ffmpeg/model/vmaf_4k_v0.6.1.json'
                                    else : 
                                        model ='ffmpeg/model/vmaf_v0.6.1.json'
                                    vmaftxt = os.path.join('vmaf.txt')
                                    if video_or.split('.')[-1] == 'yuv':
                                        if c == 'VVenC' : 
                                            #print(devid)
                                            os.system(devid+" > /dev/null 2>&1")
                                            if pxl_fmt == 'yuv420p':
                                                command1 =  'ffmpeg -f rawvideo -pix_fmt yuv420p10le -s:v '+str(res)+' -r '+str(fps)+' -i '+str(output1_yuv)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -frames:v '+str(nbr_frame)+' -lavfi "'+flt+'libvmaf" -f null /dev/null 2>&1 | grep "VMAF score:" | cut -d ":" -f2 > '+vmaftxt
                                            else:
                                                command1 =  'ffmpeg -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(output1_yuv)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -frames:v '+str(nbr_frame)+' -lavfi "'+flt+'libvmaf" -f null /dev/null 2>&1 | grep "VMAF score:" | cut -d ":" -f2 > '+vmaftxt
                                            
                                        else:
                                            command1 =  'ffmpeg -i '+str(output1)+' -f rawvideo -pix_fmt '+str(pxl_fmt)+' -s:v '+str(res)+' -r '+str(fps)+' -i '+str(input1)+' -frames:v '+str(nbr_frame)+' -lavfi "'+flt+'libvmaf" -f null /dev/null 2>&1 | grep "VMAF score:" | cut -d ":" -f2 > '+vmaftxt
                                    else :
                                        command1 =  'ffmpeg -i '+str(output1)+' -i '+str(input1)+' -lavfi "'+flt+'libvmaf" -f null /dev/null  2>&1 | grep "VMAF score:" | cut -d ":" -f2 > '+vmaftxt                             
                                    #print(command1)
                                    os.system(command1)
                                    a = open(vmaftxt, "r")
                                    vvmaf = float(a.readlines()[0].rstrip())
                                    sheet2.cell(row = idv, column = psnr_col+8).value = float(vvmaf) 
                        except:
                            print(c) 
                            vpsnr = 0.0
                            vssim = 0.0
                            vvmaf = 0.0                   
                        core = "|{:^60}|{:^10.4}| {:^11} | {:^10.10f} |{:^13.9f}|{:^8.2f}|{:^8.2f}|{:^8.2f}|"
                        if pathV == None:
                            if (ico == 1 and iqp != 0) or (ico == 0 and iqp == 0):
                                ser="ser"
                            else:
                                print(core.format(vname, float(tim), bit, float(enrg), float(co), float(vpsnr), float(vssim) ,float(vvmaf)))
                                print("".join(line1))
                        else:
                            print(core.format(vname, float(tim), bit, float(enrg), float(co), float(vpsnr), float(vssim) ,float(vvmaf)))
                            print("".join(line1))
                        wb2.save(self.results)			            
                    pasCodec = pasCodec + 8
            else:
                core = "|{:^60}|{:^80}|"
                video_ortxt = fg.red +"{:^60}".format(video_or)+ fg.rs
                texterr = fg.red +"{:^80}".format("Error: raw video does not exist")+ fg.rs
                print(core.format(video_ortxt, texterr))
                print("".join(line1))
           
#.............................................................................................

ap = argparse.ArgumentParser()
ap.add_argument("-d", "--dataset_path", default = "", help="Path to dataset")
ap.add_argument("-r", "--results", help="Path to results")
ap.add_argument("-m", "--metric", nargs='+', type=str, help='Metric quatlity')
ap.add_argument("-it", "--intervalv", nargs='+', type=float, help='interval period')
ap.add_argument("-p", "--platform", help="Path of encoding video")
ap.add_argument("-i", "--pathV", help="Path to original video")
ap.add_argument("-sp", "--speed", help="encoding speed")
ap.add_argument("-x", "--pxl_fmt", type=int, help="Pixel format of encoding video")
ap.add_argument("-s", "--resol", help="Resolution of encoding video")
ap.add_argument("-f", "--fps", type=int, help="Framerate of encoding video")
ap.add_argument("-b", "--bitrate", type=int, help="bitrate of encoding video")
ap.add_argument("-c", "--codec", type=str, help="codec of encoding video")
ap.add_argument("-o", "--outputVideo", type=str, help="output Video")
args = vars(ap.parse_args())
if args["pathV"] != None:
    q = [1, 2]
else:
    q = [2, 28]
if args["pathV"] != None:
    wb0 = ""
    wb1 = ""
else:
    if args["dataset_path"] == '':
        print(fg.red +"Error: please provide the path to the original dataset.\n"+fg.rs)
        exit()
    bitrate_target = os.path.dirname(args["dataset_path"])+"/data/bitrate-final.xlsx"
    video_info = os.path.dirname(args["dataset_path"])+"/data/video_information.xlsx"
    if os.path.exists(bitrate_target)== True:
        wb0 = load_workbook(bitrate_target)
    else:
        wb0 = ""
        print(fg.red +'\nError: bitrate-final.xlsx file of target bitrate not exist\n'+fg.rs)
        exit()
    if os.path.exists(video_info)== True:
        wb1 = load_workbook(video_info)
    else:
        wb0 = ""
        print(fg.red +'\nError: video_information.xlsx file of video information not exist\n'+fg.rs)
        exit()
if args["results"] == None:
    args["results"] = ""
if os.path.exists(args["results"])== True:
    wb2 = load_workbook(args["results"])
    sheet2 = wb2.worksheets[0]
    row_count = sheet2.max_row
else:
    print(fg.red +'\nError: {} file of encoding results not exist\n'.format(os.path.basename(args["results"]))+fg.rs)
    exit()
print("")
if  args['platform'] not in ['CPU', 'GPU']:
    print(fg.red +'Error: please select the platform type from this list {}'.format(['CPU', 'GPU'])+fg.rs)
    exit() 
sheet2.cell(row = 1, column = 1).value = 'Encoded videos'
sheet2.cell(row = 1, column = 2).value = 'Encoding time (s)'  
sheet2.cell(row = 1, column = 3).value = 'Bitrate (kb/s)'
if  args['platform'] == 'CPU':
    sheet2.cell(row = 1, column = 4).value = 'CPU utilization (%)'
    sheet2.cell(row = 1, column = 5).value = 'Memory utilization (%)'
    sheet2.cell(row = 1, column = 6).value = 'Energy (Wh)'
    sheet2.cell(row = 1, column = 7).value = 'CO2 emissions in France (g)'
    sheet2.cell(row = 1, column = 8).value = 'Equivalent of CO2 emissions in distance travelled by car (m)'
    mc = 0
    psnr_col = 9
elif  args['platform'] == 'GPU':
    sheet2.cell(row = 1, column = 4).value = 'CPU utilization (%)'
    sheet2.cell(row = 1, column = 5).value = 'Memory utilization (%)'
    sheet2.cell(row = 1, column = 6).value = 'GPU utilization (%)' 
    sheet2.cell(row = 1, column = 7).value = 'GPU Memory utilization (%)'
    sheet2.cell(row = 1, column = 8).value = 'Temperature (°C)' 
    sheet2.cell(row = 1, column = 9).value = 'Energy GPU (Wh)'
    sheet2.cell(row = 1, column = 10).value = 'CO2 emissions in France (g)'
    sheet2.cell(row = 1, column = 11).value = 'Equivalent of CO2 emissions in distance travelled by car (m)'
    mc = 0
    psnr_col = 12
if args['metric'] == None:
    args['metric'] = []
if 'psnr' in args['metric']:
    sheet2.cell(row = 1, column = psnr_col).value = 'PSNR Y'
    sheet2.cell(row = 1, column = psnr_col+1).value = 'PSNR U'
    sheet2.cell(row = 1, column = psnr_col+2).value = 'PSNR V'
    sheet2.cell(row = 1, column = psnr_col+3).value = 'YUV-PSNR'
    mc = 1
if 'ssim' in args['metric']:
    sheet2.cell(row = 1, column = psnr_col+4).value = 'SSIM Y'
    sheet2.cell(row = 1, column = psnr_col+5).value = 'SSIM U'
    sheet2.cell(row = 1, column = psnr_col+6).value = 'SSIM V'
    sheet2.cell(row = 1, column = psnr_col+7).value = 'YUV-SSIM'
    mc = 2
if 'vmaf' in args['metric']:
    sheet2.cell(row = 1, column = psnr_col+8).value = 'VMAF'
    mc = 3
wb2.save(args["results"]) 
projPath = os.path.dirname(__file__)
p1 = video_encoding(args['results'], args['dataset_path'], projPath)
if args['pxl_fmt'] == 10:
    pxl_fmtp = 'yuv420p10le'
elif args['pxl_fmt'] == 8:
    pxl_fmtp = 'yuv420p'
else:
    pxl_fmtp = ""

p1.encoding(args['metric'], q, args['platform'], wb0, wb1, wb2, args['pathV'], pxl_fmtp, args['bitrate'], args['resol'], args['fps'], args['codec'], args['speed'], args['outputVideo'], args['intervalv'])







